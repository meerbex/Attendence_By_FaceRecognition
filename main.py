# importing packages that we need
import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# if xlsx does not exists then create
if os.path.exists('AI_Class_Attendance.xlsx'):
    workbook = load_workbook(filename="AI_Class_Attendance.xlsx")
else:
    workbook = Workbook()
sheet = workbook.active
today = datetime.now().date().strftime("%d.%m.%Y")

# if tab sheet is not today, then create new with today's date
if sheet.title != today:
    sheet = workbook.create_sheet(today)
    workbook.active = sheet
    sheet[f"A1"] = "Name"
    sheet[f"B1"] = "Time"
workbook.save(filename="AI_Class_Attendance.xlsx")

# getting images from folders with their names
path = 'student_images'
images = []
classNames = []
myList = os.listdir(path)
for cl in myList:
    curImg = cv2.imread(f'{path}/{cl}')
    images.append(curImg)
    classNames.append(os.path.splitext(cl)[0])
print(classNames)


def findEncodings(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)[0]
        encodeList.append(encode)
    return encodeList


def markAttendance(name):
    found = False
    max_str_len = 0
    sheet.column_dimensions['A'].width = 20

    for i, cell in enumerate(sheet['A']):
        if cell.value == name:
            found = True
        if cell.value is not None:
            if max_str_len < len(cell.value):
                max_str_len = len(cell.value)

    if not found:
        now = datetime.now()
        dtString = now.strftime('%H:%M:%S')
        sheet[f"A{len(sheet['A']) + 1}"] = name
        sheet[f"B{len(sheet['A'])}"] = dtString
        print(f"A{len(sheet['A']) + 1}", name, dtString)


        os.system(f'say welcome {name.lower()}')

        sheet.column_dimensions['A'].width = max_str_len
        workbook.save(filename="AI_Class_Attendance.xlsx")

encodeListKnown = findEncodings(images)

cap = cv2.VideoCapture(0)

while True:
    success, img = cap.read()
    # img = captureScreen()
    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

    facesCurFrame = face_recognition.face_locations(imgS)
    encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
        # print(faceDis)
        matchIndex = np.argmin(faceDis)

        if faceDis[matchIndex] < 0.50:
            name = classNames[matchIndex].upper()
            markAttendance(name)
        else:
            name = 'Unknown'
        y1, x2, y2, x1 = faceLoc
        y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
        cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
        cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
        cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)

    cv2.imshow('Webcam', img)
    cv2.waitKey(1)

